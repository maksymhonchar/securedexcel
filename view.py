from typing import List, Dict, Any, Optional

from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtWidgets import (QAction, QFileDialog, QHBoxLayout, QLineEdit,
                             QMainWindow, QMenuBar, QPushButton, QTableWidget,
                             QTableWidgetItem, QTabWidget, QVBoxLayout,
                             QWidget, QLabel, QScrollArea)


class SpreadsheetView(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.init_ui()

    def init_ui(self):
        # set window characteristics
        self.setWindowTitle("Захищена база даних")
        self.setGeometry(100, 100, 800, 600)

        # create menu bar
        menu_bar = QMenuBar()
        file_menu = menu_bar.addMenu("File")
        edit_menu = menu_bar.addMenu("Edit")

        # set menu bar
        self.setMenuBar(menu_bar)

        # create open file action
        self.open_file_action = QAction("Open", self)
        file_menu.addAction(self.open_file_action)

        # create save file action
        self.save_file_action = QAction("Save", self)
        file_menu.addAction(self.save_file_action)

        # create new column action
        self.add_column_action = QAction("Add Column", self)
        edit_menu.addAction(self.add_column_action)

        # create search bar
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0, 0, 0, 0)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Пошук...")
        search_layout.addWidget(self.search_edit)

        self.search_button = QPushButton("Знайти")
        search_layout.addWidget(self.search_button)

        search_bar = QWidget()
        search_bar.setLayout(search_layout)

        # create results label
        self.search_results_label = QLabel('')
        self.search_results_label.setWordWrap(True)
        search_results_scroll_area = QScrollArea()
        search_results_scroll_area.setWidgetResizable(True)
        search_results_scroll_area.setWidget(self.search_results_label)
        search_results_scroll_area.setMaximumHeight(300)

        # create sheets tabs
        self.tab_widget = QTabWidget(self)
        self.setCentralWidget(self.tab_widget)

        # create main layout
        main_layout = QVBoxLayout()
        main_layout.addWidget(search_bar)
        main_layout.addWidget(search_results_scroll_area)
        main_layout.addWidget(self.tab_widget)

        # add central widget
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def get_open_file_path(self):
        file_dialog = QFileDialog()
        return file_dialog.getOpenFileName(
            parent=self,
            caption="Open Excel File",
            directory="",
            filter="XLSX Files (*.xlsx);;XLS Files (*.xls)"
        )[0]

    def get_save_file_path(self):
        file_dialog = QFileDialog()
        return file_dialog.getSaveFileName(
            parent=self,
            caption="Save Excel File",
            directory="",
            filter="XLSX Files (*.xlsx)"
        )[0]

    def get_current_search_query(self):
        return self.search_edit.text()

    def set_search_results(self, text: str) -> None:
        self.search_results_label.setText(text)

    @staticmethod
    def get_excel_style_column_name(col) -> str:
        LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        result: List[str] = []
        while col:
            col, rem = divmod(col - 1, 26)
            result[:0] = LETTERS[rem]
        return ''.join(result)

    def add_tab(self, worksheet: Worksheet) -> None:
        # add a table to display the sheet data
        table_widget = QTableWidget(self)
        table_widget.setColumnCount(worksheet.max_column)
        table_widget.setRowCount(worksheet.max_row)
        table_widget.setHorizontalHeaderLabels(
            [
                self.get_excel_style_column_name(column_idx + 1)
                for column_idx in range(worksheet.max_column)
            ]
        )
        table_widget.setVerticalHeaderLabels(
            [
                str(row_idx + 1)
                for row_idx in range(worksheet.max_row)
            ]
        )
        # set column widths and row heights based on worksheet dimensions
        for col_idx, (col_name, col_dim) in enumerate(worksheet.column_dimensions.items(), start=1):
            width = col_dim.width
            if width:
                table_widget.setColumnWidth(col_idx-1, int(width * 7))
        for row_idx, row_dim in worksheet.row_dimensions.items():
            height = row_dim.height
            if height:
                table_widget.setRowHeight(row_idx-1, int(height * 2))
        # add data to table
        for row_idx, row in enumerate(worksheet.iter_rows()):
            for col_idx, cell in enumerate(row):
                table_item = QTableWidgetItem(str(cell.value))
                table_widget.setItem(row_idx, col_idx, table_item)
        # add new tab
        sheet_title = worksheet.title
        self.tab_widget.addTab(table_widget, sheet_title)

    def get_current_tab_text(self):
        return self.tab_widget.tabText(
            self.tab_widget.currentIndex()
        )

    def get_cell_value_in_current_tab(self, row_idx: int, col_idx: int) -> Optional[Any]:
        current_tab_index = self.tab_widget.currentIndex()
        table_widget = self.tab_widget.widget(current_tab_index)
        item = table_widget.item(row_idx, col_idx)
        return item.text() if item is not None else None

    def add_column_in_current_tab(self) -> None:
        current_tab_index = self.tab_widget.currentIndex()
        table_widget = self.tab_widget.widget(current_tab_index)
        # insert new column after the last column
        num_cols = table_widget.columnCount()
        table_widget.insertColumn(num_cols)
        # set header label for the new column
        new_column_label = self.get_excel_style_column_name(num_cols + 1)
        table_widget.setHorizontalHeaderItem(num_cols, QTableWidgetItem(new_column_label))
