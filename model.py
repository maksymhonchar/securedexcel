from collections import OrderedDict, defaultdict
from copy import copy
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class SpreadsheetModel(object):

    def __init__(self) -> None:
        self.worksheets: Dict[str, Worksheet] = OrderedDict()

    def load_spreadsheet(self, file_path: str) -> None:
        self.worksheets = OrderedDict()
        workbook = openpyxl.load_workbook(file_path)
        for worksheet in workbook.worksheets:
            worksheet_title = worksheet.title
            self.worksheets[worksheet_title] = worksheet

    @staticmethod
    def _get_cell_with_nonempty_fill(worksheet: Worksheet) -> Cell:
        attempts_cells: List[str] = [
            "A1", "B1", "C1",
            "A2", "B2", "C2",
            "A3", "B3", "C3",
        ]
        for cell_location in attempts_cells:
            cell = worksheet[cell_location]
            if cell.fill.patternType is None:
                continue
            else:
                return cell
        # return "default" cell        
        default_ecll_location = attempts_cells[0]
        return worksheet[default_ecll_location]

    def save_spreadsheet(self, file_path: str) -> None:
        workbook = openpyxl.Workbook()
        for worksheet_title, worksheet in self.worksheets.items():
            new_worksheet = workbook.create_sheet(title=worksheet_title)
            # copy cell formatting
            cell_with_nonempty_fill = self._get_cell_with_nonempty_fill(worksheet)
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    new_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    new_cell = new_worksheet.cell(row=row_idx, column=col_idx)
                    new_cell.number_format = cell_with_nonempty_fill.number_format
                    new_cell.font = copy(cell_with_nonempty_fill.font)
                    new_cell.border = copy(cell_with_nonempty_fill.border)
                    new_cell.fill = copy(cell_with_nonempty_fill.fill)
                    new_cell.alignment = copy(cell_with_nonempty_fill.alignment)
            # copy worksheet attributes
            new_worksheet.sheet_format = copy(worksheet.sheet_format)
            new_worksheet.sheet_properties = copy(worksheet.sheet_properties)
            new_worksheet.merged_cells = copy(worksheet.merged_cells)
            new_worksheet.page_margins = copy(worksheet.page_margins)
            new_worksheet.freeze_panes = copy(worksheet.freeze_panes)
            # transfer row and column dimensions
            for col_letter, col_dim in worksheet.column_dimensions.items():
                new_worksheet.column_dimensions[col_letter] = copy(col_dim)
            for row_num, row_dim in worksheet.row_dimensions.items():
                new_worksheet.row_dimensions[row_num] = copy(row_dim)
        # remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(worksheet=workbook['Sheet'])
        # save the workbook
        workbook.save(file_path)

    def get_worksheets_titles(self) -> List[str]:
        return list(self.worksheets.keys())

    def get_worksheet(self, sheet_title: str) -> Worksheet:
        return self.worksheets[sheet_title]

    def update_cell(self, sheet_title: str, row_idx: int, col_idx: int, value: Any) -> None:
        worksheet = self.worksheets[sheet_title]
        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
        cell.value = value

    def add_column(self, sheet_title: str) -> None:
        worksheet = self.worksheets[sheet_title]
        worksheet.insert_cols(worksheet.max_column + 1)

    def search(self, query: str) -> Dict[str, List[Any]]:
        search_results: Dict[str, List[Any]] = defaultdict(list)
        for sheet_title, worksheet in self.worksheets.items():
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and query.lower() in cell.value.lower():
                        search_results[sheet_title].append(cell.value)
                    elif isinstance(cell.value, (int, float)) and str(cell.value) == query:
                        search_results[sheet_title].append(cell.value)
        return search_results
